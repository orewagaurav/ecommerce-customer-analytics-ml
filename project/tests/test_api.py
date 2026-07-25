"""API contract tests.

Exercised through FastAPI's TestClient against a real container, so routing,
dependency wiring, serialisation and exception handling are all covered. The
app is built with `create_app`, which is why no module-level singleton needs
patching.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from src.api.main import create_app
from src.config import Settings

PROJECT_ROOT = Path(__file__).resolve().parents[1]
MODELS_DIR = PROJECT_ROOT / "models"
FEATURE_STORE = PROJECT_ROOT / "feature_store" / "features.parquet"

pytestmark = pytest.mark.skipif(
    not (FEATURE_STORE.exists() and (MODELS_DIR / "churn_model_artifacts.joblib").exists()),
    reason="feature store or model artifacts not built",
)


@pytest.fixture(scope="module")
def settings(tmp_path_factory) -> Settings:
    """Point history at a temp dir so tests never touch the real audit log."""
    tmp = tmp_path_factory.mktemp("api")
    return Settings(
        feature_store_path=FEATURE_STORE,
        model_path=MODELS_DIR,
        registry_path=MODELS_DIR / "model_registry.json",
        history_path=tmp / "history.parquet",
        log_format="console",
        log_level="WARNING",
    )


@pytest.fixture(scope="module")
def client(settings) -> TestClient:
    with TestClient(create_app(settings)) as test_client:
        yield test_client


@pytest.fixture(scope="module")
def known_customer_id(client) -> int:
    return client.get("/v1/customers?limit=1").json()["customer_ids"][0]


# --- health -----------------------------------------------------------------

def test_health_reports_healthy_when_dependencies_are_present(client):
    body = client.get("/v1/health").json()

    assert body["status"] == "healthy"
    assert body["models_loaded"] is True
    assert body["feature_store_available"] is True


def test_root_advertises_docs_and_health(client):
    body = client.get("/").json()
    assert body["docs"] == "/docs"


def test_openapi_schema_is_served(client):
    schema = client.get("/openapi.json").json()
    assert "/v1/predict/{customer_id}" in schema["paths"]


# --- predict ----------------------------------------------------------------

def test_predict_returns_full_contract(client, known_customer_id):
    response = client.post(f"/v1/predict/{known_customer_id}")
    assert response.status_code == 200

    body = response.json()
    for key in (
        "customer_id", "cluster_label", "predicted_clv", "churn_probability",
        "decision", "recommendation_actions", "explanations", "shap_top_features",
        "model_version", "latency_ms",
    ):
        assert key in body


def test_predict_churn_probability_is_bounded(client, known_customer_id):
    body = client.post(f"/v1/predict/{known_customer_id}").json()
    assert 0.0 <= body["churn_probability"] <= 1.0


def test_predict_returns_nonzero_shap(client, known_customer_id):
    """Regression guard for the zero-background explainer bug, at API level."""
    body = client.post(f"/v1/predict/{known_customer_id}").json()

    for task in ("CLV", "Churn"):
        contributions = [abs(f["contribution"]) for f in body["shap_top_features"][task]]
        assert sum(contributions) > 0


def test_predict_can_skip_explanations(client, known_customer_id):
    body = client.post(
        f"/v1/predict/{known_customer_id}", json={"include_explanations": False}
    ).json()

    assert body["shap_top_features"]["CLV"] == []
    assert body["predicted_clv"] is not None


def test_predict_unknown_customer_returns_404_envelope(client):
    response = client.post("/v1/predict/99999999")

    assert response.status_code == 404
    body = response.json()
    assert body["error"] == "customer_not_found"
    assert body["request_id"]


def test_predict_rejects_negative_customer_id(client):
    assert client.post("/v1/predict/-5").status_code == 422


def test_response_carries_request_id_and_timing_headers(client, known_customer_id):
    response = client.post(f"/v1/predict/{known_customer_id}")

    assert response.headers["X-Request-ID"]
    assert float(response.headers["X-Response-Time-Ms"]) >= 0


def test_supplied_request_id_is_echoed(client, known_customer_id):
    response = client.post(
        f"/v1/predict/{known_customer_id}", headers={"X-Request-ID": "trace-abc"}
    )
    assert response.headers["X-Request-ID"] == "trace-abc"


# --- model-info -------------------------------------------------------------

def test_model_info_exposes_registry(client):
    body = client.get("/v1/model-info").json()

    assert body["production_version"].startswith("v")
    tasks = {model["task"] for model in body["models"]}
    assert {"clv", "churn", "segmentation"} <= tasks


def test_model_info_includes_feature_list_and_metrics(client):
    models = client.get("/v1/model-info").json()["models"]
    churn = next(model for model in models if model["task"] == "churn")

    assert "PredictedCLV" in churn["feature_list"]
    assert "ROC_AUC" in churn["metrics"]


def test_model_info_reports_feature_store_stats(client):
    store = client.get("/v1/model-info").json()["feature_store"]

    assert store["customer_count"] > 0
    assert store["source_rows"] > 0


# --- metrics ----------------------------------------------------------------

def test_metrics_counts_requests_and_predictions(client, known_customer_id):
    before = client.get("/v1/metrics").json()["process"]
    client.post(f"/v1/predict/{known_customer_id}")
    after = client.get("/v1/metrics").json()["process"]

    assert after["predictions"] == before["predictions"] + 1
    assert after["requests"] > before["requests"]


def test_metrics_reports_uptime_and_model_version(client):
    body = client.get("/v1/metrics").json()

    assert body["process"]["uptime_seconds"] >= 0
    assert body["model_version"].startswith("v")


def test_metrics_latency_populated_after_predictions(client, known_customer_id):
    client.post(f"/v1/predict/{known_customer_id}")
    lifetime = client.get("/v1/metrics").json()["lifetime"]

    assert lifetime["avg_latency_ms"] is not None
    assert lifetime["avg_latency_ms"] > 0


def test_lifetime_predictions_reflect_persisted_history(client, known_customer_id):
    """Lifetime counters come from the audit log, so a restart cannot zero them
    while latency aggregates computed from the same rows stay populated."""
    client.post(f"/v1/predict/{known_customer_id}")
    lifetime = client.get("/v1/metrics").json()["lifetime"]

    assert lifetime["predictions"] > 0
    assert lifetime["unique_customers"] >= 1


# --- simulate ---------------------------------------------------------------

def test_simulate_returns_baseline_and_simulated(client, known_customer_id):
    response = client.post(
        f"/v1/simulate/{known_customer_id}", json={"overrides": {"Frequency": 50.0}}
    )
    assert response.status_code == 200

    body = response.json()
    assert body["baseline"]["cluster_label"]
    assert body["simulated"]["cluster_label"]
    assert body["applied_overrides"] == {"Frequency": 50.0}


def test_simulate_changes_the_prediction(client, known_customer_id):
    """A large behavioural override must actually move the model output."""
    body = client.post(
        f"/v1/simulate/{known_customer_id}",
        json={"overrides": {"Frequency": 200.0, "Monetary": 250000.0, "Recency": 1.0}},
    ).json()

    moved = (
        abs(body["clv_delta"]) > 0
        or abs(body["churn_delta"]) > 0
        or body["segment_changed"]
    )
    assert moved, "overriding every RFM dimension left the prediction unchanged"


def test_simulate_rejects_unknown_feature(client, known_customer_id):
    response = client.post(
        f"/v1/simulate/{known_customer_id}", json={"overrides": {"NotAFeature": 1.0}}
    )
    assert response.status_code == 422


def test_simulate_unknown_customer_returns_404(client):
    response = client.post("/v1/simulate/99999999", json={"overrides": {"Frequency": 5.0}})
    assert response.status_code == 404


def test_simulate_does_not_write_to_history(client, known_customer_id):
    """A hypothetical is not a prediction the system made, so it must not be logged."""
    before = client.get("/v1/metrics").json()["lifetime"]["predictions"]
    client.post(f"/v1/simulate/{known_customer_id}", json={"overrides": {"Frequency": 9.0}})
    after = client.get("/v1/metrics").json()["lifetime"]["predictions"]

    assert after == before


# --- history ----------------------------------------------------------------

def test_history_returns_recorded_predictions(client, known_customer_id):
    client.post(f"/v1/predict/{known_customer_id}")
    body = client.get("/v1/history").json()

    assert body["total"] > 0
    assert body["entries"][0]["customer_id"] is not None


def test_history_filters_by_customer(client, known_customer_id):
    client.post(f"/v1/predict/{known_customer_id}")
    body = client.get(f"/v1/history?customer_id={known_customer_id}").json()

    assert all(entry["customer_id"] == known_customer_id for entry in body["entries"])


def test_history_filters_by_churn_probability(client, known_customer_id):
    client.post(f"/v1/predict/{known_customer_id}")
    body = client.get("/v1/history?min_churn_probability=0.99").json()

    assert all(entry["churn_probability"] >= 0.99 for entry in body["entries"])


def test_history_respects_limit(client, known_customer_id):
    for _ in range(3):
        client.post(f"/v1/predict/{known_customer_id}")
    body = client.get("/v1/history?limit=2").json()

    assert body["returned"] <= 2


def test_history_rejects_invalid_limit(client):
    assert client.get("/v1/history?limit=0").status_code == 422


# --- customer profile -------------------------------------------------------

def test_customer_profile_returns_features(client, known_customer_id):
    body = client.get(f"/v1/customers/{known_customer_id}/profile").json()

    assert body["customer_id"] == known_customer_id
    for feature in ("Recency", "Frequency", "Monetary", "Tenure"):
        assert feature in body["features"]


def test_customer_profile_unknown_customer_returns_404(client):
    assert client.get("/v1/customers/99999999/profile").status_code == 404


# --- reports ----------------------------------------------------------------

def test_customer_pdf_report_returns_a_pdf(client, known_customer_id):
    response = client.get(f"/v1/reports/customer/{known_customer_id}/pdf")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    # %PDF magic bytes: proves a real document, not an error page with a
    # convincing content-type.
    assert response.content[:4] == b"%PDF"
    assert len(response.content) > 2000


def test_customer_pdf_sets_download_filename(client, known_customer_id):
    response = client.get(f"/v1/reports/customer/{known_customer_id}/pdf")
    assert f"customer_{known_customer_id}_report.pdf" in response.headers["content-disposition"]


def test_customer_pdf_unknown_customer_returns_404(client):
    assert client.get("/v1/reports/customer/99999999/pdf").status_code == 404


def test_customers_excel_workbook_is_valid(client):
    import io
    import openpyxl

    response = client.get("/v1/reports/customers/excel?limit=50")
    assert response.status_code == 200
    assert response.content[:2] == b"PK"  # xlsx is a zip archive

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    for sheet in ("Summary", "Customer Features", "Monthly Revenue", "Country Revenue"):
        assert sheet in workbook.sheetnames


def test_customers_excel_respects_limit(client):
    import io
    import openpyxl

    response = client.get("/v1/reports/customers/excel?limit=25")
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    # +1 for the header row.
    assert workbook["Customer Features"].max_row <= 26


def test_history_excel_workbook_is_valid(client, known_customer_id):
    import io
    import openpyxl

    client.post(f"/v1/predict/{known_customer_id}")
    response = client.get("/v1/reports/history/excel")

    assert response.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Predictions" in workbook.sheetnames


def test_report_endpoints_appear_in_openapi(client):
    paths = client.get("/openapi.json").json()["paths"]

    assert "/v1/reports/customer/{customer_id}/pdf" in paths
    assert "/v1/reports/customers/excel" in paths
